import os
import openai
import openpyxl 
from dotenv import load_dotenv
load_dotenv() 
from datetime import datetime

class Resp(object):
    pass

def summ_prompt(ab_str):
    '''
    Given abstract string, generate prompt string
    '''
    preamble = "Generate a one sentence summary of the following paragraph.  The summary should describe what research was conducted and the key conclusion."
    return preamble + "\n\nParagraph:\n" + ab_str

def app_prompt(ab_str):
    '''
    Given abstract string, generate prompt string
    '''
    preamble = "Then generate a list of two of the most relevant military and defense applications of the engineering technology described in the following paragraph.  Each item in the list in the list should be a single sentence.  The list should be formatted in markdown"
    return preamble + "\n\nParagraph:\n" + ab_str

def cat_prompt(categories, ab_str):
    preamble = """Below is a set of categories and a one paragraph research abstract.  Select the one best category to fit the following paragraph.

"""
    for c in categories:
        preamble += c + '\n'
    
    return preamble + '\nParagraph:\n' +ab_str


# Load your API key from an environment variable or secret management service
openai.api_key = os.getenv("OPENAI_API_KEY")

datafile='ThesisAbstractsMAE.xlsx'
wb = openpyxl.load_workbook(datafile)
ws = wb.active

# Store everything as list of objects
records = []
# Read rows - skip header
for row in ws.iter_rows(
        min_row = 2, max_row = ws.max_row,
        min_col=1, max_col=ws.max_column,
        values_only=True):
    resp = Resp()
    resp.name = row[0]
    resp.abstract = row[1]
    records.append(resp)


categories = ['Materials Science and Engineering',
              'Robotics and Controls Engineering',
              'Space and Aerospace Engineering',
              'Thermodynamics, Fluid Dynamics and Energy',
              'Ship Design and Naval Engineering',
              'Survivability and Weaponeering',
              'Solid Mechanics and Structure Engineering']

model = "text-davinci-003"
#records = [records[0]]
for ii in range(len(records)):
    print("Summarizing %d of %d for <%s>"%(ii+1,len(records),records[ii].name))
    psumm = summ_prompt(records[ii].abstract)
    rsumm = openai.Completion.create(model = model,
                                     prompt = psumm,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].summary = rsumm.choices[0].text

    papp = app_prompt(records[ii].abstract)
    rapp = openai.Completion.create(model = model,
                                     prompt = papp,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].applications = rapp.choices[0].text

    pcat = cat_prompt(categories, records[ii].abstract)
    rcat = openai.Completion.create(model = model,
                                     prompt = pcat,
                                     max_tokens = 200,
                                     temperature=0.6)
    records[ii].category = rcat.choices[0].text


# Generate a summary of summaries
ssumm = ''
for record in records:
    ssumm += record.summary
prompt = 'Generate a list of between two and four categories that summarize the following engineering research projects.  The research projects are described in the following paragraphs: \n\n'+ssumm

rsumm= openai.Completion.create(model = model,
                                prompt = prompt,
                                max_tokens = 400,
                                temperature=0.6)


outf = 'synopsis.md'
print('Writing output to <%s>'%outf)
f = open(outf,'w')
now = datetime.now()
f.write("""# MAE Thesis Synopsis

Generated with <%s> on %s. \n"""%(model, now.strftime('%d %b %Y %H:%M:%S')))

f.write(rsumm.choices[0].text)

# Write categoties

for c in categories:
    f.write("## " + c + "\n\n")
    for record in records:
        if record.category.find(c) > 0:
            f.write("### " + record.name + "\n")
            f.write(record.summary)
            f.write(record.applications)


f.close()
